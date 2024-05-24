import openpyxl
import asyncio

from data_sources.models import async_session_maker
from data_sources.storages.menu_repository import MenuRepository
from data_sources.storages.submenu_reposytory import SubMenuRepository
from data_sources.storages.dish_repository import DishRepository
from config import settings


async def database_entry(path_to_table):
    """Заполняет таблицы базы данных данными
    из указанной таблицы с расширением xlsx.
    """

    wb = openpyxl.load_workbook(path_to_table)
    sheet = wb.active
    count = 1
    async with async_session_maker() as session:
        for i in sheet:
            if count == 1:
                count += 1
                continue
            if sheet.cell(row=count, column=2).value:
                menu_item = sheet.cell(row=count, column=2).value
                await MenuRepository.create_menu(
                    menu_item,
                    sheet.cell(row=count, column=3).value,
                    session,
                )
            if sheet.cell(row=count, column=5).value:
                submenu = sheet.cell(row=count, column=5).value
                from_menu = await MenuRepository.get_id_by_name(menu_item, session)
                await SubMenuRepository.create_submenu(
                    submenu,
                    sheet.cell(row=count, column=6).value,
                    from_menu,
                    session,
                )

            if sheet.cell(row=count, column=8).value:
                dish = sheet.cell(row=count, column=8).value
                from_submenu = await SubMenuRepository.get_id_by_name(submenu, session)
                await DishRepository.create_dish(
                    dish,
                    sheet.cell(row=count, column=9).value,
                    sheet.cell(row=count, column=10).value,
                    from_submenu,
                    session
                )
            count += 1

asyncio.run(database_entry(settings.path_to_table))
