"""Мониторинг изменений данных таблицы.
При изменении данных указанной таблицы с расширением xlsx,
изменения данных отображаются в таблицах базы данных.
"""
from time import sleep

import openpyxl
from sqlalchemy.exc import ProgrammingError
from fastapi import HTTPException
import asyncio

from data_sources.models import async_session_maker
from data_sources.storages.menu_repository import MenuRepository
from data_sources.storages.submenu_reposytory import SubMenuRepository
from data_sources.storages.dish_repository import DishRepository
from utils.utils import MyCustomException
from config import settings, celery


def write_dict_with_table(path_to_table):
    """Возвращает словарь со структурой меню.
    В позиции меню вложены соответствующие позиции подменю,
    в позиции подменю вложены соответствующие позиции блюд.
    """

    wb = openpyxl.load_workbook(path_to_table)
    sheet = wb.active
    count = 1
    menu = {}
    for i in sheet:
        if count == 1:
            count += 1
            continue
        if sheet.cell(row=count, column=1).value:
            number_menu = sheet.cell(row=count, column=1).value
            menu_item = sheet.cell(row=count, column=2).value
            description = sheet.cell(row=count, column=3).value
            menu[number_menu] = {}
            menu[number_menu]['title'] = menu_item
            menu[number_menu]['description'] = description
            menu[number_menu][menu_item] = {}
            menu[number_menu][menu_item]['submenu'] = {}
            menu[number_menu][menu_item]['title'] = menu_item
            menu[number_menu][menu_item]['description'] = description

        if sheet.cell(row=count, column=4).value:
            number_submenu = sheet.cell(row=count, column=4).value
            submenu = sheet.cell(row=count, column=5).value
            description = sheet.cell(row=count, column=6).value
            menu[number_menu][menu_item]['submenu'][number_submenu] = {}
            menu[number_menu][menu_item]['submenu'][number_submenu]['title'] = submenu
            menu[number_menu][menu_item]['submenu'][number_submenu]['description'] = description
            menu[number_menu][menu_item]['submenu'][number_submenu][submenu] = {}
            menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'] = {}
            menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['title'] = submenu
            menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['description'] = description
            
        if sheet.cell(row=count, column=7).value:
            number_dish = sheet.cell(row=count, column=7).value
            dish = sheet.cell(row=count, column=8).value
            description = sheet.cell(row=count, column=9).value
            price = sheet.cell(row=count, column=10).value
            menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish] = {}
            menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish][dish] = {}
            menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish]['title'] = dish
            menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish]['description'] = description
            menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish]['price'] = price
            menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish][dish]['title'] = dish
            menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish][dish]['description'] = description
            menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish][dish]['price'] = price
            
        count += 1
    return menu

def write_dict_names(path_to_table):
    """Возвращает словарь с позициями меню, подменю и блюд.
    Используется при реализации мониторинга удаления
    позиций меню, подменю либо позиций блюд.
    """

    wb = openpyxl.load_workbook(path_to_table)
    sheet = wb.active
    count = 1
    menu = {}
    for i in sheet:
        if count == 1:
            count += 1
            continue
        if sheet.cell(row=count, column=1).value:
            num_menu = sheet.cell(row=count, column=1).value
            menu_item = sheet.cell(row=count, column=2).value
            menu[(num_menu,menu_item)] = 1, menu_item, num_menu

        if sheet.cell(row=count, column=4).value:
            num_submenu = sheet.cell(row=count, column=4).value
            submenu = sheet.cell(row=count, column=5).value
            menu[(num_submenu,submenu)] = 2, submenu, num_submenu, num_menu, menu_item
            
        if sheet.cell(row=count, column=7).value:
            num_dish = sheet.cell(row=count, column=7).value
            dish = sheet.cell(row=count, column=8).value
            menu[(num_dish,dish)] = 3, dish, num_dish, num_menu, num_submenu, menu_item, submenu
            
        count += 1
    return menu


async def table_monitoring(path_to_table):
    """Реализует мониторинг изменения данных
    в указанной таблице с расширением xlsx.
    При изменении данных в указанной таблице,
    изменения отображаются в таблицах базы данных.
    """

    init_dict_with_names = write_dict_names(path_to_table) # Словарь с первоначальным состоянием позиций меню, подменю, блюд.
    init_menu = write_dict_with_table(path_to_table) # Словарь с первоначальной структурой меню.
    is_deleted = False
    
    async with async_session_maker() as session:
        while True:
            wb = openpyxl.load_workbook(path_to_table)
            sheet = wb.active
            count = 1

            for i in sheet:
                if count == 1:
                    count += 1
                    continue
                try:
                    if sheet.cell(row=count, column=1).value:
                        menu_item = sheet.cell(row=count, column=2).value
                        number_menu = sheet.cell(row=count, column=1).value
                        current_description = sheet.cell(row=count, column=3).value
                        if number_menu not in init_menu:
                            init_dict_with_names[(number_menu,menu_item)] = 1, menu_item, number_menu
                            current_description = sheet.cell(row=count, column=3).value
                            init_menu[number_menu] = {}
                            init_menu[number_menu][menu_item] = {}
                            init_menu[number_menu][menu_item]['submenu'] = {}
                            init_menu[number_menu]['title'] = menu_item
                            init_menu[number_menu]['description'] = current_description
                            init_menu[number_menu][menu_item]['title'] = menu_item
                            init_menu[number_menu][menu_item]['description'] = current_description
                            await MenuRepository.create_menu(menu_item, current_description, session)
                        else:
                            init_menu_title = init_menu[number_menu]['title']
                            init_menu_description = init_menu[number_menu]['description']
                            del init_dict_with_names[(number_menu,init_menu_title)]
                            init_dict_with_names[(number_menu,menu_item)] = 1, menu_item, number_menu
                            if init_menu_title != menu_item or init_menu_description != current_description:
                                menu_id = await MenuRepository.get_id_by_name(init_menu_title, session)
                                init_menu[number_menu]['title'] = menu_item
                                init_menu[number_menu]['description'] = current_description
                                buf = init_menu[number_menu][init_menu_title]['submenu']
                                del init_menu[number_menu][init_menu_title]['submenu']
                                init_menu[number_menu][menu_item] = {}
                                init_menu[number_menu][menu_item]['submenu'] = buf
                                init_menu[number_menu][menu_item]['title'] = menu_item
                                init_menu[number_menu][menu_item]['description'] = current_description
                                await MenuRepository.update_menu(
                                    menu_item,
                                    current_description,
                                    menu_id,
                                    session,
                                )
                    if sheet.cell(row=count, column=4).value:
                        number_submenu = sheet.cell(row=count, column=4).value
                        submenu = sheet.cell(row=count, column=5).value
                        current_submenu_desription = sheet.cell(row=count, column=6).value
                        if number_submenu not in init_menu[number_menu][menu_item]['submenu']:
                            init_dict_with_names[(number_submenu, submenu)] = 2, submenu, number_submenu, number_menu, menu_item
                            from_menu = await MenuRepository.get_id_by_name(menu_item, session)
                            init_menu[number_menu][menu_item]['submenu'][number_submenu] = {}
                            init_menu[number_menu][menu_item]['submenu'][number_submenu]['title'] = submenu
                            init_menu[number_menu][menu_item]['submenu'][number_submenu]['description'] = current_submenu_desription
                            init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu] = {}
                            init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'] = {}
                            init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['title'] = submenu
                            init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['description'] = current_submenu_desription
                            await SubMenuRepository.create_submenu(
                                submenu,
                                current_submenu_desription,
                                from_menu,
                                session,
                            )
                        else:
                            init_submenu_title = init_menu[number_menu][menu_item]['submenu'][number_submenu]['title']
                            init_submenu_description = init_menu[number_menu][menu_item]['submenu'][number_submenu]['description']
                            submenu_id = await SubMenuRepository.get_id_by_name(init_submenu_title, session)
                            from_menu = await MenuRepository.get_id_by_name(menu_item, session)
                            del init_dict_with_names[(number_submenu,init_submenu_title)]
                            init_dict_with_names[(number_submenu, submenu)] = 2, submenu, number_submenu, number_menu, menu_item
                            if submenu != init_submenu_title or init_submenu_description != current_submenu_desription:
                                buf = init_menu[number_menu][menu_item]['submenu'][number_submenu][init_submenu_title]['dish']
                                del init_menu[number_menu][menu_item]['submenu'][number_submenu][init_submenu_title]['dish']
                                init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu] = {}
                                init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'] = buf
                                init_menu[number_menu][menu_item]['submenu'][number_submenu]['title'] = submenu
                                init_menu[number_menu][menu_item]['submenu'][number_submenu]['description'] = current_submenu_desription
                                init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['title'] = submenu
                                init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['description'] = current_submenu_desription
                                await SubMenuRepository.update_submenu(
                                    submenu,
                                    current_submenu_desription,
                                    from_menu,
                                    submenu_id,
                                    session,
                                )
                    if sheet.cell(row=count, column=7).value:
                        number_dish = sheet.cell(row=count, column=7).value
                        dish = sheet.cell(row=count, column=8).value
                        current_dish_description = sheet.cell(row=count, column=9).value
                        current_price = sheet.cell(row=count, column=10).value
                        if number_dish not in init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish']:
                            init_dict_with_names[(number_dish, dish)] = 3, dish, number_dish, number_menu, number_submenu, menu_item, submenu
                            from_submenu = await SubMenuRepository.get_id_by_name(submenu, session)
                            init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish] = {}
                            init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish][dish] = {}
                            init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish]['title'] = dish
                            init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish]['description'] = current_dish_description
                            init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish]['price'] = current_price
                            init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish][dish]['title'] = dish
                            init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish][dish]['description'] = current_dish_description
                            init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish][dish]['price'] = current_price
                            await DishRepository.create_dish(
                                dish,
                                current_dish_description,
                                current_price,
                                from_submenu,
                                session,
                            )
                        else:
                            init_dish_title = init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish]['title']
                            init_description = init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish]['description']
                            init_price = init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish]['price']
                            current_price = sheet.cell(row=count, column=10).value
                            from_submenu = await SubMenuRepository.get_id_by_name(submenu, session)
                            dish_id = await DishRepository.get_id_by_name(init_dish_title, session)
                            del init_dict_with_names[(number_dish,init_dish_title)]
                            init_dict_with_names[(number_dish, dish)] = 3, dish, number_dish, number_menu, number_submenu, menu_item, submenu
                            if init_dish_title != dish or init_description != current_dish_description or init_price != current_price:
                                init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish]['title'] = dish
                                init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish]['description'] = current_dish_description
                                init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish]['price'] = current_price
                                del init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish][init_dish_title]
                                init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish][dish] = {}
                                init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish][dish]['title'] = dish
                                init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish][dish]['description'] = current_dish_description
                                init_menu[number_menu][menu_item]['submenu'][number_submenu][submenu]['dish'][number_dish][dish]['price'] = current_price
                                await DishRepository.update_dish(
                                    dish,
                                    current_dish_description,
                                    current_price,
                                    from_submenu,
                                    dish_id,
                                    session,
                                )
                except (HTTPException, KeyError, ProgrammingError):
                    await session.rollback()
                    try:
                        raise MyCustomException(
                           """
                        Ошибка.
                        В таблице нет позиции подменю для соответствующей позиции блюда,
                        либо нет позиции меню для соответствующей позиции подменю.
                        В таблице не должно быть пустых строк.
                    """ ,
                        )
                    except MyCustomException as some_ex:
                        print(some_ex)
                        return False
                count += 1
            current_dict_with_names = write_dict_names(path_to_table)
            srt_dict = sorted(
                init_dict_with_names.items(),
                key=lambda position: position[1][0],
                reverse=True,
            )
            try:
                for position in srt_dict:
                    if (position[0][0],position[1][1]) not in current_dict_with_names:
                        is_deleted = True
                        entity_number = position[1][0]
                        if entity_number == 1:
                            menu_num = position[1][2]
                            del init_menu[menu_num]
                            id_item = await MenuRepository.get_id_by_name(position[1][1], session)
                            await MenuRepository.delete_menu(id_item, session)
                        elif entity_number == 2:
                            submenu_num = position[1][2]
                            menu_num = position[1][3]
                            menu_name = position[1][4]
                            del init_menu[menu_num][menu_name]['submenu'][submenu_num]
                            id_item = await SubMenuRepository.get_id_by_name(position[1][1], session)
                            await SubMenuRepository.delete_submenu(id_item, session)
                        elif entity_number == 3:
                            dish_num = position[1][2]
                            menu_num = position[1][3]
                            submenu_num = position[1][4]
                            menu_name = position[1][5]
                            submenu_name = position[1][6]
                            del init_menu[menu_num][menu_name]['submenu'][submenu_num][submenu_name]['dish'][dish_num]
                            id_item = await DishRepository.get_id_by_name(position[1][1], session)
                            await DishRepository.delete_dish(id_item, session)
                if is_deleted:
                    init_dict_with_names = write_dict_names(path_to_table)
                    is_deleted = False
            except (HTTPException, KeyError, ProgrammingError):
                await session.rollback()
                try:
                    raise MyCustomException(
                        """
                    Ошибка.
                    В таблице нет позиции подменю для соответствующей позиции блюда,
                    либо нет позиции меню для соответствующей позиции подменю.
                    В таблице не должно быть пустых строк.
                """ ,
                    )
                except MyCustomException as some_ex:
                    print(some_ex)
                    return False    
            sleep(int(settings.check_interval))

    
@celery.task
def task_table_monitoring():
    asyncio.run(table_monitoring(settings.path_to_table))

task_table_monitoring.delay()
